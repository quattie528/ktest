#!/usr/bin/python

### MODULES ###
import datetime as dt
import openpyxl as xl #, pillow
import os
import re
import sys
import shutil
import pprint
#
import xz
#import kbench

### VARIABLES ###
debug = True
notice = True
blanklimit = 2

abc2num = {}
for i,x in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ'):
	abc2num[x] = i+1

#style of each cells
#http://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl/35560546#35560546

DRITT = True
DRITT = False

##### VIELHEIT ###############
def xls2tbls( datei, shs=[], strict=True ):
	"""
(1) BENCH with quellen-base-0629.xlsx @ 2016-08-18 17:30:26
[01] use_iterators : 2.3251   100.0   100.3   105.8   155.5
[02] options        : 2.3331    99.7   100.0   105.4   155.0
[03] read_only      : 2.4598    94.5    94.8   100.0   147.0
[04] data_only      : 3.6163    64.3    64.5    68.0   100.0

(2)
read_only : off --> 4.0746 secs
read_only : on  --> more than 60 secs (gave up to bench it)
	"""
	### VORBEREITUNG ###
#	bench = kbench.stoppuhr()
	options = {
		'data_only' : True,
		'use_iterators'  : True,
		'read_only' : True
	}
	zahl = 0
	ganz = False

	### VARIABLES ###
	try:
		wb = xl.load_workbook(datei, data_only=True, use_iterators=True, read_only=True)
	except TypeError: # rescue line on 2016-10-13 # TypeError: load_workbook() got an unexpected keyword argument 'use_iterators'
#		print( datei ) #d
		wb = xl.load_workbook(datei, data_only=True, read_only=True)
	blatten = wb.get_sheet_names()
	alle_tafelen = []
	#
	if shs == []:
		shs = blatten
		ganz = True
		alle_tafelen = {}

	### HAUPT ###
	for sh in shs:
		if isinstance(sh, int):
			sh = blatten[sh]
		try:
			ws = wb[sh]
		except KeyError:
			print( 'Keine Blatt : %s' % sh )
			if strict == True:
				ws = wb[sh]
			else:
				if ganz == True:
					alle_tafelen[sh] = []
				elif ganz == False:
					alle_tafelen.append([])
				continue
		tafel = []
		xi = ws.max_column
		yj = ws.max_row
#		print( xi,yj )

		### HAUPT ###
		for j,row in enumerate( ws.iter_rows() ):

			## Prufung ##
			# 1)
			if j >  yj: break
			#
			# 2) 2016-05-27 # x[0] -> str(x)[0]
			chk = str( row[0].value )
			if chk[0] == '#' or chk[0] == '!':
				continue

			## Werten ##
			lis = []
			for i,x in enumerate(row):
				try:
					v = x.value
					if v == None:
						v = ''
					elif isinstance(v, dt.datetime):
						if ( v.hour + v.minute + v.second ) == 0:
							v = dt.date(v.year,v.month,v.day)
# debug for 2016-08-18
#					else:
#						pass
#						print( v,type(v) )
#						if isinstance(v, str):
#							if v[0] == '=':
#								v = x.internal_value
#								print( x.internal_value )
				except UnicodeEncodeError as e:
					print(e,x)
				lis.append( v )

			## Leer ##
			if lis[0] == '':
				zahl += 1
				if zahl == blanklimit: # namely 2
					break
			else:
				zahl = 0
			tafel.append(lis)

		"""
		2019-02-01, created
		2019-02-02, commented out
#		pprint.pprint( tafel )
		if DRITT == False:
			while 1: # 2016-10-30
				if tafel == []: break
				if tafel[-1][0] == '':
					tafel.pop()
				else:
					break
#		print( '-'*40 ) #d
#		pprint.pprint( tafel )
#		print( '%'*40 ) #d
		"""

		## Enfin ##
		if ganz == True:
			alle_tafelen[sh] = tafel
		elif ganz == False:
			alle_tafelen.append(tafel)
#	wb.close()

	### AUSGABE ###
	return alle_tafelen

### EXCEL to TABLE #####
def xls2tbl( datei, sh=0, passen=0 ):
	res = xls2tbls( datei, [sh] )
	lis = []
	for x in res[0][0]:
		if isinstance(x, str):
			x = x.replace("\n",' ')
		lis.append(x)
	res[0][0] = lis
	#
	res = res[0]
	for i in range(passen): res.pop(0)
	return res

### EXCEL to DICT ###
def xls2ldic( datei, sh ):
	res = xls2tbls( datei, [sh] )
	return xz.tbl2ldic(res[0])

### EXCEL to TSVs ###
def xls2tsvs( datei,ordner ): # gigi
	res = xls2tbls(datei)
	sdatei = os.path.basename(datei)
	tdatei = sdatei.replace('.xlsx','')
	eigenpfad = ordner + '%s_%s.tsv'
	for blatt,datum in res.items():
		pfad = ordner + tdatei + '_' + blatt + '.tsv'
		pfad = pfad.replace(' ','_')
		xz.tbl2txt(datum,pfad)
	return res

"""
### XLS and XLS ###
def xls_and_xls(x1,x2,sh,*nums):
	wb1 = xl.load_workbook(x1)
	wb2 = xl.load_workbook(x2)
	ws1 = wb1[sh]
	ws2 = wb2[sh]
	#
	xi = len( ws1.columns[0] )
	for i in range(xi):
		for j in nums:
			v1 = ws1.cell(row=i+1,column=j).value
			v2 = ws2.cell(row=i+1,column=j).value
			if v1 == v2: continue
			ws1.cell(row=i+1,column=j).value = v2
	wb1.save('a.xlsx')
"""

### XLS and TABLE ###
def xls8tbl(xls,sh,tbl):
	wb = xl.load_workbook(xls)
	ws = wb[sh]
	#
	xi = len( ws.columns[0] ) #+ 1
	for list in tbl:
		xi += 1
		j = 0
		for x in list:
			j += 1
			ws.cell(row=xi,column=j).value = x
	#
	ex = 'a.xlsx'
	wb.save(ex)

	if notice == False:
		sys.stdout.write("Ausgabe als %s \n" % ex)

#

"""
##########################
### TABLE to NEW EXCEL ###
##########################
def tbl2newxls(tbl,xls): # this rarely happens
	wb = xl.Workbook()
	ws = wb.active
#	ws = wb.create_sheet()
	for i,lis in enumerate(tbl):
		for j,x in enumerate(lis):
			ws.cell(row=i+1,column=j+1).value = x
	wb.save(xls)
	if notice == False:
		sys.stdout.write("Ausgabe als %s \n" % ex)
"""

#

#########################
### DICT-TABLE to XLS ###
#########################
def dtbl2xls(dtbl,xls='a.xlsx'):
	# Schlussel muss "blatt/position" sein
	wb = xl.load_workbook(xls)

	### HAUPT ###
	for sh,tbl in dtbl.items():
		sh = sh.split('/') # blatt/position
		if len(sh) == 1:
			sh = sh[0]
			pos = 'A1'
		elif len(sh) > 1:
			pos = sh[1]
			sh = sh[0]

		ws = wb[sh]
		pos = cell2pos(pos)
		xi = pos[0]
		yj = pos[1]
		tbl = xz.bless(tbl)
		#
		for i,lis in enumerate(tbl):
			i2 = i + xi
			i2 = i + yj
			for j,x in enumerate(lis):
				j2 = j + yj
				j2 = j + xi
	#			fmt = ws.cell(row=i2,column=j2).style
				ws.cell(row=i2,column=j2).value = x

	### AUSGABE ###
	wb.save(xls)
	if notice == False:
		sys.stdout.write("Ausgabe als %s \n" % ex)

#

######################
### TABLE to EXCEL ###
######################
def tbl2xls(tbl,xls='a.xlsx',sh='Sheet1',pos='A1'):
	res = {}
	res[sh+'/'+pos] = tbl
	dtbl2xls(res,xls)

#

######################
### EXCEL to IMAGE ###
######################
def xls2img(xls,blatt):
	pass
	print( 'NOW it is yet impossible' )
	#WEB at 2015-06-09
	#http://stackoverflow.com/questions/27855317/openpyxl-can-i-create-a-horizontal-bar-chart-with-openpyxl

#

########################
### CELL to POSITION ###
########################
def cell2pos(cell):
	m = re.match('^([A-Z]{1,3})(\d+)$',cell)
	assert m
	x = 0
	w = m.group(1)
	y = int(m.group(2))

	if len(w) == 1:
		x = abc2num[w]
	elif len(w) == 2:
		x1 = int(abc2num[ w[0] ]) * 26
		x2 = int(abc2num[ w[1] ])
		x = x1 + x2
	elif len(w) == 3:
		x1 = int(abc2num[ w[0] ]) * 26 * 26
		x2 = int(abc2num[ w[1] ]) * 26
		x3 = int(abc2num[ w[2] ])
		x = x1 + x2 + x3
	return (x,y)
	return [x,y]

def test4cell2pos():
	paar = xz.txt2tbl('benchbuch/xx_pruf4cell2pos.tsv')
	for cell,antwort in paar:
		x = cell2pos(cell)
		antwort = int(antwort)
		print( cell, '->', x, '==', antwort )
		assert x[0] == antwort

def test(datei):
	wb = xl.load_workbook(datei, data_only=True, read_only=True)
	x = wb.get_sheet_names()
	print( x, type(x) )

#

##################
### XLS zu BIN ###
##################
def stamm2obj(ur,bin,blatt=0,opt={}):
#def xls2obj(bin,xls,blatt=0): # 2019-11-04
	"""
	STAMM kommst aus "STAMMDATEN"
	https://de.wikipedia.org/wiki/Stammdaten
	"""

	import xf

	### VORBEREITUNG ###
	ldic = opt.get('ldic',False)
	ddic = opt.get('ddic',False)
	ddickey = opt.get('ddickey','')
	if ddic == True:
		ldic = True
#	print( ldic ) #d
#	print( ddic ) #d
#	print( ddickey ) #d

	### HAUPT ###
	#
	## Wierderverwerten ##
	if os.path.exists(bin):
		print( '[UR  ]', xf.mtime(ur),  '||', ur  ) #d
		print( '[NACH]', xf.mtime(bin), '||', bin ) #d
#		if xf.cdate(ur) <= xf.cdate(bin):
		if xf.mtime(ur) <= xf.mtime(bin):
			print( '#Wierderverwerten' ) #d
			obj = xz.bin2obj(bin)
			return obj
	#
	## Laden ##
	print( '#Laden' ) #d
	obj = xls2tbl(ur,blatt)
	if ldic == True:
		obj = xz.tbl2ldic(obj)
		if ddic == True:
			obj = xz.ldic2ddic(obj,ddickey)
	xz.obj2bin(obj,bin)

	### AUSGABE ###
	return obj

#

#############
### MAKRO ###
#############
def xls2makro(xls,warten2=30):

	### MODULE ###
	import pyautogui as pgui
	import clipboard
	import time
	pgui.FAILSAFE = True

	### KONSTANT ###
	xls = xls.replace('/','\\')
	warten1 = 5
	
	### BEDIGUNG ###
#	os.system('taskkill /IM "excel.exe" /F ')

	### HAUPT ###
	pgui.keyDown('winleft')
	pgui.press('r')
	pgui.keyUp('winleft')
#	pgui.keyUp('windows')
#	pgui.typewrite('excel '+xls) # ASCII Keyboard
	clipboard.copy('excel '+xls)
	pgui.keyDown('ctrl')
	pgui.press('v')
	pgui.keyUp('ctrl')
	pgui.press('enter')
	#
	time.sleep(warten1)
	for i in range(5):
		pgui.press('enter')
#	xu.sleep(2)  # ich möchte Freiheit haben!
	pgui.keyDown('alt')
	pgui.press('space')
	pgui.keyUp('alt')
	pgui.press('r')
	time.sleep(1)
	#
	pgui.keyDown('alt')
	pgui.press('f8')
	pgui.keyUp('alt')
	pgui.press('enter')
#	xu.sleep(traum)  # ich möchte Freiheit haben!
	if warten2 > 0:
		time.sleep(warten2)
		#
#		pgui.keyDown('alt') # 2020-02-02
#		pgui.press('f4')    # 2020-02-02
#		pgui.keyUp('alt')   # 2020-02-02
	return True
	#
	cmd = 'taskkill /IM "excel.exe" /F '
	pgui.keyDown('winleft')
	pgui.press('r')
	pgui.keyUp('winleft')
	pgui.press('enter')
	time.sleep(1)

#

##### DIREKT ###############
if __name__=='__main__':
	pass

#TypeError: 'generator' object is not subscriptable
